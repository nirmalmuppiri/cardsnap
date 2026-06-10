import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Keys that should appear first if present, in this order
PRIORITY_KEYS = [
    "first_name", "last_name", "full_name",
    "title", "company",
    "email", "phone", "mobile",
    "website", "linkedin",
    "address",
]


def _sort_keys(all_keys: set[str]) -> list[str]:
    ordered = [k for k in PRIORITY_KEYS if k in all_keys]
    rest = sorted(all_keys - set(ordered))
    return ordered + rest


def generate_excel(contacts: list[dict], event_name: str) -> bytes:
    """
    contacts: list of dicts with keys 'id', 'event_name', 'data', 'created_at'
    Returns .xlsx file as bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = event_name[:31]  # Excel sheet name limit

    all_keys: set[str] = set()
    for c in contacts:
        all_keys.update(c["data"].keys())

    columns = _sort_keys(all_keys)

    # Header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    for col_idx, key in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=key.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, contact in enumerate(contacts, start=2):
        for col_idx, key in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=contact["data"].get(key, ""))

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
